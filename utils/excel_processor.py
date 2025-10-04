"""
Procesador de archivos Excel del banco - Versión corregida
Colegio Cristiano Macedonia
"""

import openpyxl
import xlrd
import os
from datetime import datetime
from models import db, Estudiante, Pago, CargaArchivo, ErrorProcesamiento
from werkzeug.utils import secure_filename

class ExcelProcessor:
    """Procesador de archivos Excel del banco"""
    
    def __init__(self, upload_folder='static/uploads'):
        self.upload_folder = upload_folder
        self.allowed_extensions = {'xlsx', 'xls'}
        self.usuario_id = None  # Se asignará cuando se procese
        self.resultados = {
            'archivo': '',
            'total_filas': 0,
            'procesados': 0,
            'exitosos': 0,
            'errores': 0,
            'errores_detalle': [],
            'duplicados': 0,
            'nuevos_pagos': []
        }
    
    def es_archivo_valido(self, filename):
        """Verificar si el archivo tiene extensión válida"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in self.allowed_extensions
    
    def guardar_archivo(self, archivo):
        """Guardar archivo subido de forma segura"""
        if not os.path.exists(self.upload_folder):
            os.makedirs(self.upload_folder)
        
        # Generar nombre único con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(archivo.filename)
        nombre_base, extension = os.path.splitext(filename)
        nuevo_nombre = f"{nombre_base}_{timestamp}{extension}"
        
        ruta_archivo = os.path.join(self.upload_folder, nuevo_nombre)
        archivo.save(ruta_archivo)
        
        return ruta_archivo, nuevo_nombre
    
    def leer_excel_banco(self, ruta_archivo):
        """Leer y validar archivo Excel del banco (.xls, .xlsx o TSV)"""
        try:
            extension = os.path.splitext(ruta_archivo)[1].lower()
            
            # Primero intentar detectar si es realmente un archivo Excel o TSV
            if extension == '.xls':
                # Verificar si es un archivo Excel real o TSV con extensión .xls
                try:
                    with open(ruta_archivo, 'rb') as f:
                        primeros_bytes = f.read(8)
                    
                    # Si no empieza con los bytes de Excel, es probablemente TSV
                    if not primeros_bytes.startswith(b'\xd0\xcf\x11\xe0') and not primeros_bytes.startswith(b'\x09\x08'):
                        print("🔍 Archivo .xls detectado como TSV (texto delimitado)")
                        return self._leer_tsv(ruta_archivo)
                    else:
                        return self._leer_xls(ruta_archivo)
                except:
                    # Si falla, intentar como TSV
                    print("🔍 Fallback: Leyendo como TSV")
                    return self._leer_tsv(ruta_archivo)
            elif extension == '.xlsx':
                return self._leer_xlsx(ruta_archivo)
            else:
                raise ValueError(f"Formato de archivo no soportado: {extension}")
            
        except Exception as e:
            raise ValueError(f"Error leyendo archivo Excel: {str(e)}")
    
    def _leer_tsv(self, ruta_archivo):
        """Leer archivo TSV (delimitado por tabulaciones)"""
        import csv
        
        datos = []
        
        with open(ruta_archivo, 'r', encoding='utf-8', newline='') as archivo:
            # Leer como TSV (delimitado por tabs)
            reader = csv.reader(archivo, delimiter='\t')
            
            # Leer encabezados
            headers = next(reader)
            headers = [str(header).strip().upper() for header in headers]
            
            print(f"📋 Headers TSV encontrados: {headers}")
            
            # Verificar columnas requeridas
            mapeo_columnas = self._mapear_columnas(headers)
            
            # Leer datos
            for row_num, row in enumerate(reader, start=2):
                if not any(str(cell).strip() for cell in row):  # Saltar filas vacías
                    continue
                
                fila_datos = self._procesar_fila_datos(row, headers, mapeo_columnas, row_num)
                datos.append(fila_datos)
        
        print(f"📊 TSV procesado: {len(datos)} filas de datos")
        return datos
    
    def _leer_xlsx(self, ruta_archivo):
        """Leer archivo .xlsx con openpyxl"""
        workbook = openpyxl.load_workbook(ruta_archivo)
        sheet = workbook.active
        
        # Obtener datos como lista de diccionarios
        datos = []
        headers = []
        
        # Leer encabezados (primera fila)
        for cell in sheet[1]:
            headers.append(str(cell.value).strip().upper() if cell.value else '')
        
        # Verificar columnas requeridas
        mapeo_columnas = self._mapear_columnas(headers)
        
        # Leer datos (desde la fila 2)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Saltar filas vacías
                continue
            
            fila_datos = self._procesar_fila_datos(row, headers, mapeo_columnas, row_num)
            datos.append(fila_datos)
        
        workbook.close()
        return datos
    
    def _leer_xls(self, ruta_archivo):
        """Leer archivo .xls con xlrd"""
        workbook = xlrd.open_workbook(ruta_archivo)
        sheet = workbook.sheet_by_index(0)  # Primera hoja
        
        # Obtener datos como lista de diccionarios
        datos = []
        
        # Leer encabezados (primera fila)
        headers = []
        for col in range(sheet.ncols):
            cell_value = sheet.cell_value(0, col)
            headers.append(str(cell_value).strip().upper() if cell_value else '')
        
        print(f"📋 Headers encontrados: {headers}")
        
        # Verificar columnas requeridas
        mapeo_columnas = self._mapear_columnas(headers)
        
        # Leer datos (desde la fila 1, ya que xlrd es 0-indexed)
        for row_num in range(1, sheet.nrows):
            row_data = []
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(row_num, col)
                row_data.append(cell_value)
            
            # Saltar filas vacías
            if not any(str(cell).strip() for cell in row_data if cell):
                continue
            
            fila_datos = self._procesar_fila_datos(row_data, headers, mapeo_columnas, row_num + 1)
            datos.append(fila_datos)
        
        return datos
    
    def _mapear_columnas(self, headers):
        """Mapear columnas requeridas"""
        columnas_requeridas = ['CARNET', 'NOMBRE', 'MES PAGO', 'TOTAL']
        mapeo_columnas = {}
        
        for col_req in columnas_requeridas:
            encontrada = False
            for i, header in enumerate(headers):
                if col_req in header:
                    mapeo_columnas[col_req] = i
                    encontrada = True
                    break
            
            if not encontrada:
                raise ValueError(f"Columna requerida '{col_req}' no encontrada")
        
        return mapeo_columnas
    
    def _procesar_fila_datos(self, row, headers, mapeo_columnas, row_num):
        """Procesar una fila de datos y crear diccionario"""
        fila_datos = {}
        
        # Mapear columnas requeridas
        for col_name, col_index in mapeo_columnas.items():
            if col_index < len(row):
                fila_datos[col_name] = row[col_index]
            else:
                fila_datos[col_name] = None
        
        # Agregar otras columnas opcionales
        columnas_opcionales = ['BOLETA', 'INSCRIP.', 'CUOTA', 'UTILES', 'BUS', 
                             'EXAMENES', 'BONO', 'SEGURO', 'CURSOS', 'OTROS', 
                             'MORA', 'EFECTIVO', 'CH.PROPIOS', 'CH.LOCALES', 'AGENCIA PAGO']
        
        for col_opc in columnas_opcionales:
            for i, header in enumerate(headers):
                if col_opc in header and i < len(row):
                    fila_datos[col_opc] = row[i]
                    break
            
            if col_opc not in fila_datos:
                fila_datos[col_opc] = 0 if col_opc != 'AGENCIA PAGO' else ''
        
        fila_datos['_row_num'] = row_num
        return fila_datos
    
    def procesar_archivo_banco(self, ruta_archivo, usuario_id):
        """Procesar archivo completo del banco"""
        self.usuario_id = usuario_id  # Guardar usuario_id para uso posterior
        
        try:
            # Leer archivo
            datos = self.leer_excel_banco(ruta_archivo)
            
            # Inicializar resultados
            self.resultados['archivo'] = os.path.basename(ruta_archivo)
            self.resultados['total_filas'] = len(datos)
            
            # Crear registro de carga
            carga = CargaArchivo(
                nombre_archivo=self.resultados['archivo'],
                usuario_id=usuario_id,
                registros_procesados=len(datos)
            )
            db.session.add(carga)
            db.session.flush()  # Para obtener el ID
            
            # Procesar cada fila
            for fila in datos:
                try:
                    self._procesar_fila_pago(fila, fila['_row_num'], carga.id)
                    self.resultados['procesados'] += 1
                    
                except Exception as e:
                    self._registrar_error(fila, fila['_row_num'], str(e), carga.id)
                    self.resultados['errores'] += 1
            
            # Actualizar estadísticas de carga
            carga.registros_exitosos = self.resultados['exitosos']
            carga.registros_fallidos = self.resultados['errores']
            carga.observaciones = f"Procesados: {self.resultados['exitosos']}, Errores: {self.resultados['errores']}, Duplicados: {self.resultados['duplicados']}"
            
            db.session.commit()
            
            return True, self.resultados
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error en procesamiento: {e}")
            return False, {'error': str(e)}
    
    def _procesar_fila_pago(self, fila, numero_fila, carga_id):
        """Procesar una fila individual de pago"""
        # Validar y convertir carnet
        try:
            carnet = int(float(str(fila['CARNET']))) if fila['CARNET'] else None
        except (ValueError, TypeError):
            raise ValueError(f"Carnet inválido: {fila['CARNET']}")
        
        if not carnet:
            raise ValueError("Carnet es requerido")
        
        mes_pago = str(fila['MES PAGO']).strip() if fila['MES PAGO'] else ''
        if not mes_pago:
            raise ValueError("Mes de pago es requerido")
        
        # Convertir total
        try:
            total_pagado = float(fila['TOTAL']) if fila['TOTAL'] else 0
        except (ValueError, TypeError):
            raise ValueError(f"Total inválido: {fila['TOTAL']}")
        
        # Buscar estudiante
        estudiante = Estudiante.query.filter_by(carnet=carnet, activo=True).first()
        if not estudiante:
            raise ValueError(f"Estudiante con carnet {carnet} no encontrado")
        
        # Parsear mes y año
        mes, anio = self._parsear_mes_anio(mes_pago)
        
        # Verificar si ya existe el pago
        pago_existente = Pago.query.filter_by(
            estudiante_id=estudiante.id,
            mes=mes,
            anio=anio
        ).first()
        
        if pago_existente:
            self.resultados['duplicados'] += 1
            # Traducir mes a español para el mensaje
            from utils.meses import traducir_mes
            mes_es = traducir_mes(mes, a_espanol=True)
            raise ValueError(f"Pago ya existe para {mes_es} {anio}")
        
        # Función para convertir montos
        def convertir_monto(valor):
            if valor is None or valor == '':
                return 0
            try:
                # Manejar formatos como "280.00", "0.00", etc.
                valor_str = str(valor).replace(',', '').replace('Q', '').strip()
                return float(valor_str)
            except (ValueError, TypeError):
                return 0
        
        # Crear nuevo pago
        nuevo_pago = Pago(
            estudiante_id=estudiante.id,
            mes=mes,
            anio=anio,
            fecha_pago=datetime.now().date(),
            boleta=str(fila.get('BOLETA', '')),
            
            # Conceptos de pago
            inscripcion=convertir_monto(fila.get('INSCRIP.')),
            cuota=convertir_monto(fila.get('CUOTA')),
            utiles=convertir_monto(fila.get('UTILES')),
            bus=convertir_monto(fila.get('BUS')),
            examenes=convertir_monto(fila.get('EXAMENES')),
            bono=convertir_monto(fila.get('BONO')),
            seguro=convertir_monto(fila.get('SEGURO')),
            cursos=convertir_monto(fila.get('CURSOS')),
            otros=convertir_monto(fila.get('OTROS')),
            mora=convertir_monto(fila.get('MORA')),
            
            total_pagado=total_pagado,
            
            # Métodos de pago
            efectivo=convertir_monto(fila.get('EFECTIVO')),
            cheque_propios=convertir_monto(fila.get('CH.PROPIOS')),
            cheque_locales=convertir_monto(fila.get('CH.LOCALES')),
            agencia_pago=str(fila.get('AGENCIA PAGO', '')),
            
            # CORRECCIÓN: usar self.usuario_id en lugar de variable no definida
            procesado_por=self.usuario_id
        )
        
        db.session.add(nuevo_pago)
        
        # Registrar pago exitoso
        self.resultados['exitosos'] += 1
        self.resultados['nuevos_pagos'].append({
            'carnet': carnet,
            'nombre': estudiante.nombre,
            'mes': mes,
            'anio': anio,
            'total': total_pagado
        })
    
    def _parsear_mes_anio(self, mes_pago):
        """Parsear string de mes y año del banco - CORREGIDO para fechas Excel"""
        try:
            # Limpiar el valor
            if mes_pago is None:
                mes_pago = ''
            
            mes_pago_str = str(mes_pago).strip()
            
            # Si es una fecha datetime de Excel (formato: 2025-08-01 00:00:00)
            if '-' in mes_pago_str and ('00:00:00' in mes_pago_str or len(mes_pago_str) == 10):
                try:
                    from datetime import datetime
                    # Parsear como fecha
                    if '00:00:00' in mes_pago_str:
                        fecha = datetime.strptime(mes_pago_str.split(' ')[0], '%Y-%m-%d')
                    else:
                        fecha = datetime.strptime(mes_pago_str, '%Y-%m-%d')
                    
                    # Convertir a nombre de mes en inglés
                    mes = fecha.strftime('%B')  # 'August', 'September', etc.
                    anio = fecha.year
                    
                    print(f"✅ Fecha parseada correctamente: {mes_pago_str} -> {mes} {anio}")
                    return mes, anio
                except Exception as e:
                    print(f"⚠️ Error parseando fecha Excel '{mes_pago_str}': {e}")
            
            # Formato del banco: "08/2025", "07/2025", etc.
            if '/' in mes_pago_str:
                partes = mes_pago_str.split('/')
                if len(partes) == 2:
                    mes_num = int(partes[0])
                    anio = int(partes[1])
                    
                    # Convertir número de mes a nombre en inglés
                    meses = [
                        '', 'January', 'February', 'March', 'April', 'May', 'June',
                        'July', 'August', 'September', 'October', 'November', 'December'
                    ]
                    
                    if 1 <= mes_num <= 12:
                        mes = meses[mes_num]
                        return mes, anio
            
            # Intentar formato "Enero 2025", "Jan 2025", etc. (fallback)
            partes = mes_pago_str.split()
            if len(partes) >= 2:
                mes_str = partes[0].lower()
                anio_str = partes[1]
                
                # Mapeo de meses
                meses_map = {
                    'enero': 'January', 'febrero': 'February', 'marzo': 'March',
                    'abril': 'April', 'mayo': 'May', 'junio': 'June',
                    'julio': 'July', 'agosto': 'August', 'septiembre': 'September',
                    'octubre': 'October', 'noviembre': 'November', 'diciembre': 'December',
                    'jan': 'January', 'feb': 'February', 'mar': 'March',
                    'apr': 'April', 'may': 'May', 'jun': 'June',
                    'jul': 'July', 'aug': 'August', 'sep': 'September',
                    'oct': 'October', 'nov': 'November', 'dec': 'December'
                }
                
                mes = meses_map.get(mes_str, mes_str.title())
                anio = int(anio_str)
                
                return mes, anio
            
            # Si no puede parsear, usar mes y año actual como fallback
            ahora = datetime.now()
            print(f"⚠️ No se pudo parsear '{mes_pago_str}', usando mes actual")
            return ahora.strftime('%B'), ahora.year
            
        except Exception as e:
            # Fallback a mes actual
            from datetime import datetime
            ahora = datetime.now()
            print(f"❌ Error parseando '{mes_pago}': {e}, usando mes actual")
            return ahora.strftime('%B'), ahora.year
    
    def _registrar_error(self, fila, numero_fila, error_descripcion, carga_id):
        """Registrar error de procesamiento"""
        carnet = fila.get('CARNET', 'N/A')
        nombre = fila.get('NOMBRE', 'N/A')
        
        # Convertir datos de fila a JSON serializable
        datos_fila_json = {}
        for key, value in fila.items():
            if hasattr(value, 'strftime'):  # Es una fecha
                datos_fila_json[key] = value.strftime('%Y-%m-%d %H:%M:%S')
            else:
                datos_fila_json[key] = value
        
        error = ErrorProcesamiento(
            carga_id=carga_id,
            fila_excel=numero_fila,
            carnet_estudiante=int(carnet) if str(carnet).isdigit() else None,
            error_descripcion=error_descripcion,
            datos_fila=datos_fila_json  # Usar versión serializable
        )
        
        db.session.add(error)
        
        # Agregar a resultados para mostrar al usuario
        self.resultados['errores_detalle'].append({
            'fila': numero_fila,
            'carnet': carnet,
            'nombre': nombre,
            'error': error_descripcion
        })
    
    def validar_archivo_estructura(self, ruta_archivo):
        """Validar estructura del archivo antes de procesar"""
        try:
            extension = os.path.splitext(ruta_archivo)[1].lower()
            print(f"🔍 Validando archivo {extension}")
            
            if extension == '.xlsx':
                return self._validar_xlsx(ruta_archivo)
            elif extension == '.xls':
                # Detectar si es Excel real o TSV
                try:
                    with open(ruta_archivo, 'rb') as f:
                        primeros_bytes = f.read(8)
                    
                    if not primeros_bytes.startswith(b'\xd0\xcf\x11\xe0') and not primeros_bytes.startswith(b'\x09\x08'):
                        print("🔍 Archivo .xls detectado como TSV para validación")
                        return self._validar_tsv(ruta_archivo)
                    else:
                        return self._validar_xls(ruta_archivo)
                except:
                    return self._validar_tsv(ruta_archivo)
            else:
                return False, f"Formato de archivo no soportado: {extension}"
            
        except Exception as e:
            return False, f"Error validando archivo: {str(e)}"
    
    def _validar_tsv(self, ruta_archivo):
        """Validar archivo TSV (delimitado por tabulaciones)"""
        import csv
        
        try:
            with open(ruta_archivo, 'r', encoding='utf-8', newline='') as archivo:
                reader = csv.reader(archivo, delimiter='\t')
                
                # Leer encabezados
                headers = next(reader)
                headers = [str(header).strip().upper() for header in headers]
                
                print(f"🔍 Headers TSV encontrados: {headers}")
                
                # Verificar columnas mínimas
                columnas_requeridas = ['CARNET', 'NOMBRE', 'MES PAGO', 'TOTAL']
                columnas_faltantes = []
                
                for col_req in columnas_requeridas:
                    encontrada = any(col_req in header for header in headers)
                    if not encontrada:
                        columnas_faltantes.append(col_req)
                
                # Contar filas de datos
                num_filas = sum(1 for _ in reader)
                
                if num_filas < 1:
                    return False, "El archivo está vacío (no tiene datos)"
                
                # Verificar formato leyendo primera fila de datos
                archivo.seek(0)  # Volver al inicio
                reader = csv.reader(archivo, delimiter='\t')
                next(reader)  # Saltar headers
                
                try:
                    primera_fila = next(reader)
                    if len(primera_fila) >= 4:
                        mes_ejemplo = primera_fila[3]  # Columna MES PAGO
                        print(f"🔍 Formato de mes ejemplo: {mes_ejemplo}")
                except StopIteration:
                    pass
                
                if columnas_faltantes:
                    return False, f"Columnas faltantes: {', '.join(columnas_faltantes)}"
                
                print(f"✅ Estructura del archivo TSV válida ({num_filas} filas)")
                return True, f"Estructura válida - Archivo TSV del banco detectado ({num_filas} registros)"
                
        except Exception as e:
            return False, f"Error validando archivo TSV: {str(e)}"
    
    def _validar_xlsx(self, ruta_archivo):
        """Validar archivo .xlsx"""
        workbook = openpyxl.load_workbook(ruta_archivo)
        sheet = workbook.active
        
        # Leer primera fila (headers)
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).upper().strip() if cell.value else '')
        
        print(f"🔍 Headers encontrados (.xlsx): {headers}")
        
        # Verificar columnas mínimas
        columnas_requeridas = ['CARNET', 'NOMBRE', 'MES PAGO', 'TOTAL']
        columnas_faltantes = []
        
        for col_req in columnas_requeridas:
            encontrada = any(col_req in header for header in headers)
            if not encontrada:
                columnas_faltantes.append(col_req)
        
        # Verificar que tenga datos
        if sheet.max_row < 2:
            workbook.close()
            return False, "El archivo está vacío (no tiene datos)"
        
        # Verificar formato de mes en primera fila de datos
        if sheet.max_row >= 2:
            primera_fila_datos = list(sheet[2])
            if len(primera_fila_datos) >= 4:
                mes_ejemplo = primera_fila_datos[3].value
                print(f"🔍 Formato de mes ejemplo: {mes_ejemplo}")
        
        workbook.close()
        
        if columnas_faltantes:
            return False, f"Columnas faltantes: {', '.join(columnas_faltantes)}"
        
        print("✅ Estructura del archivo .xlsx válida")
        return True, "Estructura válida - Formato Excel .xlsx detectado"
    
    def _validar_xls(self, ruta_archivo):
        """Validar archivo .xls"""
        workbook = xlrd.open_workbook(ruta_archivo)
        sheet = workbook.sheet_by_index(0)
        
        # Leer primera fila (headers)
        headers = []
        for col in range(sheet.ncols):
            cell_value = sheet.cell_value(0, col)
            headers.append(str(cell_value).upper().strip() if cell_value else '')
        
        print(f"🔍 Headers encontrados (.xls): {headers}")
        
        # Verificar columnas mínimas
        columnas_requeridas = ['CARNET', 'NOMBRE', 'MES PAGO', 'TOTAL']
        columnas_faltantes = []
        
        for col_req in columnas_requeridas:
            encontrada = any(col_req in header for header in headers)
            if not encontrada:
                columnas_faltantes.append(col_req)
        
        # Verificar que tenga datos
        if sheet.nrows < 2:
            return False, "El archivo está vacío (no tiene datos)"
        
        # Verificar formato de mes en primera fila de datos
        if sheet.nrows >= 2:
            mes_ejemplo = sheet.cell_value(1, 3) if sheet.ncols >= 4 else None  # Fila 1, Col 3 (MES PAGO)
            print(f"🔍 Formato de mes ejemplo: {mes_ejemplo}")
        
        if columnas_faltantes:
            return False, f"Columnas faltantes: {', '.join(columnas_faltantes)}"
        
        print("✅ Estructura del archivo .xls válida")
        return True, "Estructura válida - Formato Excel .xls (banco) detectado"