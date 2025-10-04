"""
Generación de reportes - Módulo completo
"""
from flask import Blueprint, render_template, request, jsonify, make_response, send_file
from flask_login import login_required, current_user
from models import Estudiante, Grado, Pago, db
from datetime import datetime, date
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import and_, or_

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

@reportes_bp.route('/')
@login_required
def index():
    """Índice de reportes disponibles"""
    return render_template('reportes/index.html')

@reportes_bp.route('/por-grado')
@login_required
def por_grado():
    """Reporte de pagos por grado"""
    # Obtener parámetros
    mes = request.args.get('mes', datetime.now().strftime('%B'))
    anio = request.args.get('anio', datetime.now().year, type=int)
    grado_id = request.args.get('grado_id', type=int)
    
    # Construir query
    query = db.session.query(
        Grado.nombre.label('grado'),
        db.func.count(Estudiante.id).label('total_estudiantes'),
        db.func.count(Pago.id).label('estudiantes_al_dia'),
        (db.func.count(Estudiante.id) - db.func.count(Pago.id)).label('estudiantes_morosos')
    ).select_from(Grado)\
     .join(Estudiante, Grado.id == Estudiante.grado_id)\
     .outerjoin(Pago, and_(
         Pago.estudiante_id == Estudiante.id,
         Pago.mes == mes,
         Pago.anio == anio
     ))\
     .filter(Estudiante.activo == True, Grado.activo == True)
    
    if grado_id:
        query = query.filter(Grado.id == grado_id)
    
    resultados = query.group_by(Grado.id, Grado.nombre).order_by(Grado.id).all()
    
    # Obtener lista de grados para filtro
    grados = Grado.query.filter_by(activo=True).order_by(Grado.nombre).all()
    
    return render_template('reportes/por_grado.html',
                         resultados=resultados,
                         grados=grados,
                         mes_seleccionado=mes,
                         anio_seleccionado=anio,
                         grado_seleccionado=grado_id)

@reportes_bp.route('/morosos')
@login_required
def morosos():
    """Reporte de estudiantes morosos"""
    mes = request.args.get('mes', datetime.now().strftime('%B'))
    anio = request.args.get('anio', datetime.now().year, type=int)
    grado_id = request.args.get('grado_id', type=int)
    
    # Estudiantes sin pago del mes
    query = db.session.query(Estudiante)\
                     .join(Grado)\
                     .outerjoin(Pago, and_(
                         Pago.estudiante_id == Estudiante.id,
                         Pago.mes == mes,
                         Pago.anio == anio
                     ))\
                     .filter(
                         Estudiante.activo == True,
                         Pago.id.is_(None)  # Sin pago
                     )
    
    if grado_id:
        query = query.filter(Estudiante.grado_id == grado_id)
    
    estudiantes_morosos = query.order_by(Grado.nombre, Estudiante.nombre).all()
    
    # Obtener lista de grados para filtro
    grados = Grado.query.filter_by(activo=True).order_by(Grado.nombre).all()
    
    return render_template('reportes/morosos.html',
                         estudiantes=estudiantes_morosos,
                         grados=grados,
                         mes_seleccionado=mes,
                         anio_seleccionado=anio,
                         grado_seleccionado=grado_id)

@reportes_bp.route('/al-dia')
@login_required
def al_dia():
    """Reporte de estudiantes al día"""
    mes = request.args.get('mes', datetime.now().strftime('%B'))
    anio = request.args.get('anio', datetime.now().year, type=int)
    grado_id = request.args.get('grado_id', type=int)
    
    # Estudiantes con pago del mes
    query = db.session.query(Estudiante, Pago)\
                     .join(Grado)\
                     .join(Pago, and_(
                         Pago.estudiante_id == Estudiante.id,
                         Pago.mes == mes,
                         Pago.anio == anio
                     ))\
                     .filter(Estudiante.activo == True)
    
    if grado_id:
        query = query.filter(Estudiante.grado_id == grado_id)
    
    estudiantes_al_dia = query.order_by(Grado.nombre, Estudiante.nombre).all()
    
    # Obtener lista de grados para filtro
    grados = Grado.query.filter_by(activo=True).order_by(Grado.nombre).all()
    
    return render_template('reportes/al_dia.html',
                         estudiantes=estudiantes_al_dia,
                         grados=grados,
                         mes_seleccionado=mes,
                         anio_seleccionado=anio,
                         grado_seleccionado=grado_id)

@reportes_bp.route('/exportar/morosos')
@login_required
def exportar_morosos():
    """Exportar lista de morosos a Excel"""
    mes = request.args.get('mes', datetime.now().strftime('%B'))
    anio = request.args.get('anio', datetime.now().year, type=int)
    grado_id = request.args.get('grado_id', type=int)
    
    # Obtener datos
    query = db.session.query(Estudiante)\
                     .join(Grado)\
                     .outerjoin(Pago, and_(
                         Pago.estudiante_id == Estudiante.id,
                         Pago.mes == mes,
                         Pago.anio == anio
                     ))\
                     .filter(
                         Estudiante.activo == True,
                         Pago.id.is_(None)
                     )
    
    if grado_id:
        query = query.filter(Estudiante.grado_id == grado_id)
    
    estudiantes = query.order_by(Grado.nombre, Estudiante.nombre).all()
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes Morosos"
    
    # Encabezados
    headers = ['Carnet', 'Nombre', 'Grado', 'Sección', 'Cuota', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Datos
    for row, estudiante in enumerate(estudiantes, 2):
        ws.cell(row=row, column=1, value=estudiante.carnet)
        ws.cell(row=row, column=2, value=estudiante.nombre)
        ws.cell(row=row, column=3, value=estudiante.grado.nombre)
        ws.cell(row=row, column=4, value=estudiante.seccion or '')
        ws.cell(row=row, column=5, value=float(estudiante.cuota_aplicable))
        ws.cell(row=row, column=6, value=estudiante.observaciones or '')
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    grado_filtro = f"_grado_{grado_id}" if grado_id else ""
    filename = f"morosos_{mes}_{anio}{grado_filtro}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@reportes_bp.route('/exportar/al-dia')
@login_required
def exportar_al_dia():
    """Exportar lista de estudiantes al día a Excel"""
    mes = request.args.get('mes', datetime.now().strftime('%B'))
    anio = request.args.get('anio', datetime.now().year, type=int)
    grado_id = request.args.get('grado_id', type=int)
    
    # Obtener datos
    query = db.session.query(Estudiante, Pago)\
                     .join(Grado)\
                     .join(Pago, and_(
                         Pago.estudiante_id == Estudiante.id,
                         Pago.mes == mes,
                         Pago.anio == anio
                     ))\
                     .filter(Estudiante.activo == True)
    
    if grado_id:
        query = query.filter(Estudiante.grado_id == grado_id)
    
    resultados = query.order_by(Grado.nombre, Estudiante.nombre).all()
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes Al Día"
    
    # Encabezados
    headers = ['Carnet', 'Nombre', 'Grado', 'Sección', 'Total Pagado', 'Fecha Pago', 'Boleta']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Datos
    for row, (estudiante, pago) in enumerate(resultados, 2):
        ws.cell(row=row, column=1, value=estudiante.carnet)
        ws.cell(row=row, column=2, value=estudiante.nombre)
        ws.cell(row=row, column=3, value=estudiante.grado.nombre)
        ws.cell(row=row, column=4, value=estudiante.seccion or '')
        ws.cell(row=row, column=5, value=float(pago.total_pagado))
        ws.cell(row=row, column=6, value=pago.fecha_pago.strftime('%d/%m/%Y') if pago.fecha_pago else '')
        ws.cell(row=row, column=7, value=pago.boleta or '')
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    grado_filtro = f"_grado_{grado_id}" if grado_id else ""
    filename = f"al_dia_{mes}_{anio}{grado_filtro}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@reportes_bp.route('/exportar/resumen-grados')
@login_required
def exportar_resumen_grados():
    """Exportar resumen por grados a Excel"""
    mes = request.args.get('mes', datetime.now().strftime('%B'))
    anio = request.args.get('anio', datetime.now().year, type=int)
    
    # Obtener datos
    resultados = db.session.query(
        Grado.nombre.label('grado'),
        db.func.count(Estudiante.id).label('total_estudiantes'),
        db.func.count(Pago.id).label('estudiantes_al_dia'),
        (db.func.count(Estudiante.id) - db.func.count(Pago.id)).label('estudiantes_morosos'),
        (Grado.cuota_mensual * db.func.count(Pago.id)).label('recaudacion_estimada')
    ).select_from(Grado)\
     .join(Estudiante, Grado.id == Estudiante.grado_id)\
     .outerjoin(Pago, and_(
         Pago.estudiante_id == Estudiante.id,
         Pago.mes == mes,
         Pago.anio == anio
     ))\
     .filter(Estudiante.activo == True, Grado.activo == True)\
     .group_by(Grado.id, Grado.nombre, Grado.cuota_mensual)\
     .order_by(Grado.id).all()
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Resumen {mes} {anio}"
    
    # Título
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value=f"RESUMEN DE PAGOS - {mes.upper()} {anio}")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Grado', 'Total Estudiantes', 'Al Día', 'Morosos', '% Al Día', 'Recaudación Estimada']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Datos
    total_estudiantes = 0
    total_al_dia = 0
    total_recaudacion = 0
    
    for row, resultado in enumerate(resultados, 4):
        ws.cell(row=row, column=1, value=resultado.grado)
        ws.cell(row=row, column=2, value=resultado.total_estudiantes)
        ws.cell(row=row, column=3, value=resultado.estudiantes_al_dia)
        ws.cell(row=row, column=4, value=resultado.estudiantes_morosos)
        
        porcentaje = (resultado.estudiantes_al_dia / resultado.total_estudiantes * 100) if resultado.total_estudiantes > 0 else 0
        ws.cell(row=row, column=5, value=f"{porcentaje:.1f}%")
        ws.cell(row=row, column=6, value=float(resultado.recaudacion_estimada or 0))
        
        total_estudiantes += resultado.total_estudiantes
        total_al_dia += resultado.estudiantes_al_dia
        total_recaudacion += resultado.recaudacion_estimada or 0
    
    # Totales
    total_row = len(resultados) + 5
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total_estudiantes).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=total_al_dia).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_estudiantes - total_al_dia).font = Font(bold=True)
    
    porcentaje_total = (total_al_dia / total_estudiantes * 100) if total_estudiantes > 0 else 0
    ws.cell(row=total_row, column=5, value=f"{porcentaje_total:.1f}%").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_recaudacion).font = Font(bold=True)
    
    # Ajustar anchos de columna
    column_widths = [15, 15, 10, 10, 10, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    filename = f"resumen_grados_{mes}_{anio}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@reportes_bp.route('/api/meses-disponibles')
@login_required
def api_meses_disponibles():
    """API para obtener meses con pagos registrados"""
    meses_pagos = db.session.query(
        Pago.mes, 
        Pago.anio,
        db.func.count(Pago.id).label('total_pagos')
    ).group_by(Pago.mes, Pago.anio)\
     .order_by(Pago.anio.desc(), Pago.mes.desc()).all()
    
    from utils.meses import traducir_mes

    meses_disponibles = []
    for mes_pago in meses_pagos:
        mes_es = traducir_mes(mes_pago.mes, a_espanol=True)
        meses_disponibles.append({
            'mes': mes_pago.mes,
            'mes_es': mes_es,
            'anio': mes_pago.anio,
            'total_pagos': mes_pago.total_pagos,
            'display': f"{mes_es} {mes_pago.anio}"
        })
    
    return jsonify({
        'success': True,
        'meses': meses_disponibles
    })

@reportes_bp.route('/api/estadisticas-generales')
@login_required
def api_estadisticas_generales():
    """API para estadísticas generales del sistema"""
    from utils.meses import traducir_mes

    # Mes actual
    mes_actual = datetime.now().strftime('%B')
    mes_actual_es = traducir_mes(mes_actual, a_espanol=True)
    anio_actual = datetime.now().year
    
    # Estadísticas básicas
    total_estudiantes = Estudiante.query.filter_by(activo=True).count()
    total_grados = Grado.query.filter_by(activo=True).count()
    
    # Estudiantes al día este mes
    estudiantes_al_dia = db.session.query(Estudiante.id).join(Pago).filter(
        Estudiante.activo == True,
        Pago.mes == mes_actual,
        Pago.anio == anio_actual
    ).distinct().count()
    
    # Recaudación del mes
    recaudacion_mes = db.session.query(db.func.sum(Pago.total_pagado)).filter(
        Pago.mes == mes_actual,
        Pago.anio == anio_actual
    ).scalar() or 0
    
    # Últimos 6 meses de recaudación
    recaudacion_historica = db.session.query(
        Pago.mes,
        Pago.anio,
        db.func.sum(Pago.total_pagado).label('total'),
        db.func.count(Pago.id).label('num_pagos')
    ).group_by(Pago.mes, Pago.anio)\
     .order_by(Pago.anio.desc(), Pago.mes.desc())\
     .limit(6).all()
    
    return jsonify({
        'success': True,
        'estadisticas': {
            'total_estudiantes': total_estudiantes,
            'total_grados': total_grados,
            'estudiantes_al_dia': estudiantes_al_dia,
            'estudiantes_morosos': total_estudiantes - estudiantes_al_dia,
            'porcentaje_al_dia': round((estudiantes_al_dia / total_estudiantes * 100) if total_estudiantes > 0 else 0, 1),
            'recaudacion_mes': float(recaudacion_mes),
            'mes_actual': f"{mes_actual_es} {anio_actual}"
        },
        'recaudacion_historica': [
            {
                'mes': r.mes,
                'mes_es': traducir_mes(r.mes, a_espanol=True),
                'anio': r.anio,
                'total': float(r.total),
                'num_pagos': r.num_pagos,
                'display': f"{traducir_mes(r.mes, a_espanol=True)} {r.anio}"
            } for r in recaudacion_historica
        ]
    })